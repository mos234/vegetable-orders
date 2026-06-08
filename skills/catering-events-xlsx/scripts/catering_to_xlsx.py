#!/usr/bin/env python3
"""
catering_to_xlsx.py — Convert Hebrew catering PDF/DOCX to Excel
Usage: python catering_to_xlsx.py <input.pdf|docx> <output.xlsx>
"""
import sys, re
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Styles ────────────────────────────────────────────────────────────
hfont    = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hfill    = PatternFill('solid', fgColor='4472C4')
cat_font = Font(name='Arial', bold=True, size=11, color='1F4E79')
cat_fill = PatternFill('solid', fgColor='D6E4F0')
scf      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
scfill   = PatternFill('solid', fgColor='2E75B6')
tot_f    = Font(name='Arial', bold=True, size=11)
tot_fill = PatternFill('solid', fgColor='FFF2CC')
day_fill = PatternFill('solid', fgColor='375623')
note_f   = Font(name='Arial', size=10, italic=True, color='C00000')
dfont    = Font(name='Arial', size=11)
bfont    = Font(name='Arial', bold=True, size=11)
border   = Border(left=Side('thin'), right=Side('thin'),
                  top=Side('thin'), bottom=Side('thin'))
center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_al = Alignment(horizontal='right',  vertical='center', wrap_text=True)

PER_GUEST = {'סלטים', 'תוספות', 'מנות אחרונות', 'מזנון סיום', 'מזנון פתיחה'}
CAT_ORDER = ['סלטים','ראשונות','מנת פתיחה','מזנון פתיחה',
             'עיקריות','תוספות','מנות אחרונות','מזנון סיום','הערות']
CAT_KEYWORDS = set(CAT_ORDER)

# ── Quantity resolution ───────────────────────────────────────────────
def resolve_qty(qty, category, guests):
    if category not in PER_GUEST:
        return qty, ''
    if qty == 1:
        return guests, ''
    if isinstance(qty, (int, float)):
        half = guests / 2
        if abs(qty - half) / half <= 0.25:
            return qty, 'חצי מהמוזמנים'
    return qty, ''

# ── PDF text extraction ───────────────────────────────────────────────
def extract_pdf_lines(path):
    try:
        import pdfplumber
        lines = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text = page.extract_text(x_tolerance=3, y_tolerance=3)
                if text:
                    lines.extend(text.splitlines())
        return lines
    except ImportError:
        import fitz  # PyMuPDF
        doc = fitz.open(path)
        lines = []
        for page in doc:
            lines.extend(page.get_text().splitlines())
        return lines

def extract_docx_lines(path):
    from docx import Document
    doc = Document(path)
    return [p.text.strip() for p in doc.paragraphs if p.text.strip()]

# ── Event parser ──────────────────────────────────────────────────────
EVENT_HEADER_RE = re.compile(
    r'(\d{4})\s+(.+?)\s+אירוע\s+(\d+)\s+(\d+)(?:\s+(.+))?$'
)
DATE_LINE_RE = re.compile(
    r'(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})\s+(\d{2}:\d{2})\s+יום\s+([א-ת])\s+(\S+)'
)
ITEM_RE = re.compile(r'^(\d+)\s+(.+?)(?:\s+מטבח\s+(?:קר|חם))?$')

def parse_events(lines):
    events = []
    current = None
    current_cat = None

    for line in lines:
        line = line.strip()
        if not line or line.startswith('Page ') or 'טווח תאריכים' in line:
            continue

        # Event header
        m = EVENT_HEADER_RE.match(line)
        if m:
            if current:
                events.append(current)
            current = {
                'order': int(m.group(1)),
                'name': m.group(2).strip(),
                'guests': int(m.group(3)),
                'option': int(m.group(4)),
                'contact': (m.group(5) or '').strip(),
                'date': '', 'day': '', 'time': '', 'hall': '',
                'service': 'הגשה לשולחנות',
                'categories': OrderedDict(),
            }
            current_cat = None
            continue

        if current is None:
            continue

        # Date/hall line
        m = DATE_LINE_RE.match(line)
        if m:
            current['date'] = m.group(1)
            current['time'] = f"{m.group(2)}-{m.group(3)}"
            current['day']  = m.group(4)
            current['hall'] = m.group(5)
            continue

        # Category keyword
        clean = line.rstrip(':')
        if clean in CAT_KEYWORDS:
            current_cat = clean
            if current_cat not in current['categories']:
                current['categories'][current_cat] = []
            continue

        # Item line: starts with a number
        m = ITEM_RE.match(line)
        if m and current_cat:
            qty = int(m.group(1))
            name = m.group(2).strip()
            current['categories'][current_cat].append((name, qty))
            continue

        # Fallback: plain text line in current category
        if current_cat and line:
            current['categories'][current_cat].append((line, 1))

    if current:
        events.append(current)
    return events

# ── Excel helpers ─────────────────────────────────────────────────────
def sc(ws, row, col, val, font=None, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=val)
    if font:  c.font  = font
    if fill:  c.fill  = fill
    if align: c.alignment = align
    c.border = border
    return c

def write_event_sheet(wb, event, is_first=False):
    title = f"{event['hall']} - {event['name']}"[:31]
    ws = wb.active if is_first else wb.create_sheet(title)
    if is_first: ws.title = title
    ws.sheet_view.rightToLeft = True
    g = event['guests']

    ws.merge_cells('A1:E1')
    sc(ws,1,1, f"הזמנה {event['order']} | {event['name']} | אולם {event['hall']} | יום {event['day']} {event['date']} {event['time']}",
       Font(name='Arial',bold=True,size=14,color='1F4E79'), align=Alignment(horizontal='center',vertical='center'))
    ws.merge_cells('A2:E2')
    sc(ws,2,1, f"כמות אורחים: {g} | אופציה: {event['option']} | איש קשר: {event['contact']} | {event['service']}",
       Font(name='Arial',size=11,italic=True), align=Alignment(horizontal='center',vertical='center'))

    row = 4
    for cat in CAT_ORDER:
        if cat not in event['categories']: continue
        items = event['categories'][cat]
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=5)
        for col in range(1,6): ws.cell(row=row,column=col).fill=cat_fill; ws.cell(row=row,column=col).border=border
        sc(ws,row,1,cat,cat_font,cat_fill,Alignment(horizontal='right',vertical='center'))
        row += 1
        for ci,ch in enumerate(['#','פריט','כמות','הערה'],1): sc(ws,row,ci,ch,hfont,hfill,center)
        ws.cell(row=row,column=5).border=border; row += 1
        for idx,(name,qty) in enumerate(items,1):
            dq, note = resolve_qty(qty, cat, g)
            sc(ws,row,1,idx,dfont,align=center)
            sc(ws,row,2,name,dfont,align=right_al)
            sc(ws,row,3,dq,bfont if isinstance(dq,int) else dfont,align=center)
            sc(ws,row,4,note,note_f if note else dfont,align=right_al)
            ws.cell(row=row,column=5).border=border; row+=1
        row+=1

    ws.column_dimensions['A'].width=6
    ws.column_dimensions['B'].width=52
    ws.column_dimensions['C'].width=12
    ws.column_dimensions['D'].width=22

def write_summary(wb, events, title_str, insert_pos=0):
    ws = wb.create_sheet('ריכוז כללי', insert_pos)
    ws.sheet_view.rightToLeft = True
    ws.merge_cells('A1:F1')
    sc(ws,1,1,f'ריכוז כללי - {title_str}',Font(name='Arial',bold=True,size=16,color='1F4E79'),
       align=Alignment(horizontal='center',vertical='center'))
    ev_info = ' | '.join(f"{e['hall']} {e['name']} ({e['guests']})" for e in events)
    ws.merge_cells('A2:F2')
    sc(ws,2,1,ev_info,Font(name='Arial',size=10,italic=True),align=Alignment(horizontal='center',vertical='center'))

    row = 4
    for cat in CAT_ORDER:
        if cat == 'הערות': continue
        item_data = defaultdict(lambda:{'total':0,'events':[]})
        has = any(cat in e['categories'] for e in events)
        if not has: continue
        for ev in events:
            if cat not in ev['categories']: continue
            for name,qty in ev['categories'][cat]:
                dq,_ = resolve_qty(qty,cat,ev['guests'])
                if not isinstance(dq,int): continue
                item_data[name]['total'] += dq
                item_data[name]['events'].append(f"{ev['hall']}/{ev['day']} ({dq})")

        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6)
        for col in range(1,7): ws.cell(row=row,column=col).fill=scfill; ws.cell(row=row,column=col).border=border
        sc(ws,row,1,cat,scf,scfill,Alignment(horizontal='right',vertical='center'))
        row+=1
        for ci,ch in enumerate(['#','פריט','סה"כ','מס\' אירועים','פירוט'],1): sc(ws,row,ci,ch,hfont,hfill,center)
        ws.cell(row=row,column=6).border=border; row+=1

        cat_total=0
        for idx,(name,data) in enumerate(item_data.items(),1):
            sc(ws,row,1,idx,dfont,align=center)
            sc(ws,row,2,name,dfont,align=right_al)
            sc(ws,row,3,data['total'],bfont,align=center)
            sc(ws,row,4,len(data['events']),dfont,align=center)
            sc(ws,row,5,' | '.join(data['events']),dfont,align=right_al)
            ws.cell(row=row,column=6).border=border
            cat_total+=data['total']; row+=1

        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
        for col in range(1,7): ws.cell(row=row,column=col).fill=tot_fill; ws.cell(row=row,column=col).border=border
        sc(ws,row,1,f'סה"כ {cat}',tot_f,tot_fill,Alignment(horizontal='right',vertical='center'))
        sc(ws,row,3,cat_total,tot_f,tot_fill,center); row+=2

    ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=45
    ws.column_dimensions['C'].width=12; ws.column_dimensions['D'].width=14
    ws.column_dimensions['E'].width=45

def write_day_sheet(wb, day_label, date_str, day_events):
    ws = wb.create_sheet(f"יום {day_label} - {date_str.replace('/','.')}")
    ws.sheet_view.rightToLeft = True
    ws.merge_cells('A1:F1')
    sc(ws,1,1,f"יום {day_label} - {date_str}",
       Font(name='Arial',bold=True,size=14,color='FFFFFF'),day_fill,
       Alignment(horizontal='center',vertical='center'))
    row=3
    for ev in day_events:
        g=ev['guests']
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6)
        ev_fill=PatternFill('solid',fgColor='E2EFDA')
        for col in range(1,7): ws.cell(row=row,column=col).fill=ev_fill; ws.cell(row=row,column=col).border=border
        sc(ws,row,1,f"אולם {ev['hall']} | {ev['name']} | {g} אורחים",
           Font(name='Arial',bold=True,size=12,color='375623'),ev_fill,
           Alignment(horizontal='right',vertical='center'))
        row+=1
        for cat in CAT_ORDER:
            if cat not in ev['categories'] or cat=='הערות': continue
            ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=6)
            for col in range(1,7): ws.cell(row=row,column=col).fill=cat_fill; ws.cell(row=row,column=col).border=border
            sc(ws,row,2,cat,cat_font,cat_fill,Alignment(horizontal='right',vertical='center'))
            row+=1
            for name,qty in ev['categories'][cat]:
                dq,note=resolve_qty(qty,cat,g)
                ws.cell(row=row,column=1).border=border
                sc(ws,row,2,name,dfont,align=right_al)
                sc(ws,row,3,dq,bfont if isinstance(dq,int) else dfont,align=center)
                sc(ws,row,4,note,note_f if note else dfont,align=right_al)
                ws.cell(row=row,column=5).border=border; ws.cell(row=row,column=6).border=border; row+=1
        row+=1
    ws.column_dimensions['A'].width=4; ws.column_dimensions['B'].width=52
    ws.column_dimensions['C'].width=12; ws.column_dimensions['D'].width=22

# ── Main ──────────────────────────────────────────────────────────────
def build_workbook(events, output_path):
    wb = Workbook()
    for i,ev in enumerate(events):
        write_event_sheet(wb,ev,is_first=(i==0))

    days = {}
    for ev in events:
        days.setdefault((ev['day'],ev['date']),[]).append(ev)
    for (day_lbl,date_str),day_evs in days.items():
        write_day_sheet(wb,day_lbl,date_str,day_evs)

    dates = sorted(set(e['date'] for e in events))
    date_range = f"{dates[0]}-{dates[-1]}" if len(dates)>1 else dates[0]
    write_summary(wb,events,date_range,insert_pos=0)

    wb.save(output_path)
    print(f"✅ Saved {output_path} — {len(wb.sheetnames)} sheets")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python catering_to_xlsx.py <input.pdf|docx> <output.xlsx>")
        sys.exit(1)
    path = sys.argv[1]
    out  = sys.argv[2]
    if path.endswith('.pdf'):
        lines = extract_pdf_lines(path)
    else:
        lines = extract_docx_lines(path)
    events = parse_events(lines)
    print(f"Parsed {len(events)} events")
    build_workbook(events, out)
