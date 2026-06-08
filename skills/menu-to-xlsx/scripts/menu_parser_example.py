"""
Example parser script showing how to parse a Hebrew menu document.
This can be used as reference or run directly.
Usage: python menu_parser_example.py <input.docx> <output.xlsx>
"""
import sys
import re
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CATEGORY_KEYWORDS = [
    'סלטים', 'מנה ראשונה', 'תוספות', 'מנה עיקרית',
    'מנה אחרונה', 'קינוחים', 'משקאות', 'פריטים'
]

def extract_text_from_docx(path):
    doc = Document(path)
    lines = []
    for p in doc.paragraphs:
        text = p.text.strip()
        if text:
            lines.append(text)
    return lines

def parse_menu(lines):
    title = None
    halls = []
    current_hall = None
    current_category = None

    for line in lines:
        line = line.strip().replace('**', '')
        if not line:
            continue

        # Detect title (first line usually)
        if title is None and ('תפריט' in line or 'בוקר' in line or 'ערב' in line):
            title = line
            continue

        # Detect logistics lines (entry/exit)
        if 'כניסה' in line or 'יציאה' in line:
            continue

        # Detect hall line: contains "אולם" or "מנות"
        hall_match = re.search(r'(.+?)[-–\s]+(\d+)\s*מנות', line)
        if hall_match or 'אולם' in line:
            if current_hall:
                halls.append(current_hall)
            hall_name = hall_match.group(1).strip() if hall_match else line.strip()
            portions = int(hall_match.group(2)) if hall_match else None
            current_hall = {'name': hall_name, 'portions': portions, 'categories': {}}
            current_category = None
            continue

        if current_hall is None:
            continue

        # Detect category header
        is_category = False
        for kw in CATEGORY_KEYWORDS:
            if line.startswith(kw) or line.rstrip(':').strip() == kw:
                current_category = kw
                is_category = True
                # Check if items are on same line after ":"
                after = line.split(':', 1)
                if len(after) > 1 and after[1].strip():
                    items = [i.strip() for i in after[1].split(',') if i.strip()]
                    current_hall['categories'].setdefault(current_category, []).extend(items)
                break

        if not is_category and current_hall:
            if current_category is None:
                current_category = 'פריטים'
            current_hall['categories'].setdefault(current_category, []).append(line)

    if current_hall:
        halls.append(current_hall)

    return title, halls

def build_xlsx(title, halls, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "תפריט"
    ws.sheet_view.rightToLeft = True

    # Collect all categories across halls
    all_cats = []
    for h in halls:
        for c in h['categories']:
            if c not in all_cats:
                all_cats.append(c)

    headers = ['אולם', 'כמות מנות'] + all_cats
    start_row = 1

    # Title row
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws['A1'] = title
        ws['A1'].font = Font(name='Arial', bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        start_row = 2

    # Styles
    hfont = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='4472C4')
    dfont = Font(name='Arial', size=11)
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    for i, hall in enumerate(halls):
        row = start_row + 1 + i
        ws.cell(row=row, column=1, value=hall['name']).font = dfont
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=1).border = border

        ws.cell(row=row, column=2, value=hall['portions']).font = dfont
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=2).border = border

        for ci, cat in enumerate(all_cats):
            items = hall['categories'].get(cat, [])
            val = ', '.join(items) if items else ''
            cell = ws.cell(row=row, column=3 + ci, value=val)
            cell.font = dfont
            cell.alignment = center
            cell.border = border

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    for ci in range(len(all_cats)):
        col_letter = chr(ord('C') + ci)
        ws.column_dimensions[col_letter].width = max(20, min(55, 10 + max((len(', '.join(h['categories'].get(all_cats[ci], []))) for h in halls), default=0)))

    wb.save(output_path)

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python menu_parser_example.py <input.docx> <output.xlsx>")
        sys.exit(1)
    lines = extract_text_from_docx(sys.argv[1])
    title, halls = parse_menu(lines)
    build_xlsx(title, halls, sys.argv[2])
    print(f"Created {sys.argv[2]} with {len(halls)} halls")
